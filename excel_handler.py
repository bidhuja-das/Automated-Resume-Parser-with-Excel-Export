import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from datetime import datetime
import os
from config import EXCEL_FILE

def initialize_excel():
    """Create Excel file with headers if it doesn't exist"""
    if not os.path.exists(EXCEL_FILE):
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "Resume Data"
        
        # Headers
        headers = ['Sr. No.', 'Date & Time', 'File Name', 'Name', 'Email', 'Phone', 'Skills']
        sheet.append(headers)
        
        # Style headers
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=12)
        
        for cell in sheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Set column widths
        sheet.column_dimensions['A'].width = 10
        sheet.column_dimensions['B'].width = 20
        sheet.column_dimensions['C'].width = 30
        sheet.column_dimensions['D'].width = 25
        sheet.column_dimensions['E'].width = 30
        sheet.column_dimensions['F'].width = 20
        sheet.column_dimensions['G'].width = 50
        
        workbook.save(EXCEL_FILE)
        print("✅ Excel file created successfully!")
    else:
        print("📊 Excel file already exists.")

def append_to_excel(data):
    """Append parsed data to Excel file"""
    try:
        workbook = openpyxl.load_workbook(EXCEL_FILE)
        sheet = workbook.active
        
        # Get next row number
        next_row = sheet.max_row + 1
        sr_no = next_row - 1  # Subtract header row
        
        # Current date and time
        current_datetime = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        
        # Prepare row data
        row_data = [
            sr_no,
            current_datetime,
            data['filename'],
            data['name'],
            data['email'],
            data['phone'],
            data['skills']
        ]
        
        # Append to sheet
        sheet.append(row_data)
        
        # Center align all cells in the new row
        for cell in sheet[next_row]:
            cell.alignment = Alignment(horizontal='left', vertical='center')
        
        # Save workbook
        workbook.save(EXCEL_FILE)
        print(f"✅ Data added to Excel (Row {sr_no})")
        return True
        
    except Exception as e:
        print(f"❌ Error writing to Excel: {e}")
        return False

def get_excel_stats():
    """Get statistics from Excel file"""
    try:
        if not os.path.exists(EXCEL_FILE):
            return 0
        
        workbook = openpyxl.load_workbook(EXCEL_FILE)
        sheet = workbook.active
        total_records = sheet.max_row - 1  # Exclude header
        workbook.close()
        return total_records
    except:
        return 0