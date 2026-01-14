import openpyxl
from config import EXCEL_FILE

try:
    workbook = openpyxl.load_workbook(EXCEL_FILE)
    sheet = workbook.active
    
    print("📊 Current Excel Data:")
    print("=" * 80)
    
    # Print headers
    headers = [cell.value for cell in sheet[1]]
    print(f"{'Sr. No.':<8} {'Date & Time':<20} {'File Name':<25} {'Name':<20} {'Email':<25}")
    print("-" * 80)
    
    # Print data rows
    for row in range(2, sheet.max_row + 1):
        sr_no = sheet.cell(row=row, column=1).value
        date_time = sheet.cell(row=row, column=2).value
        filename = sheet.cell(row=row, column=3).value
        name = sheet.cell(row=row, column=4).value
        email = sheet.cell(row=row, column=5).value
        
        print(f"{sr_no:<8} {str(date_time):<20} {filename:<25} {name:<20} {email:<25}")
    
    workbook.close()
    
except Exception as e:
    print(f"❌ Error reading Excel: {e}")
